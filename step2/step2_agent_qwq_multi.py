import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import re
import math
import json
import openpyxl
import pandas as pd
from openpyxl import Workbook
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
import time
from openai import OpenAI
from transformers import AutoModelForCausalLM, AutoTokenizer
from prompts.step2_common_0 import return_messages
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()  
api_key = os.getenv("ALI_API_KEY")
os.environ["CUDA_DEVICE_ORDER"] = "PCI_BUS_ID"
os.environ["CUDA_VISIBLE_DEVICES"] = "5,6,7"

client = OpenAI(
    api_key=api_key,  
    base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",  # DashScope
)

# 全局线程锁，用于并发写文件时加锁，避免冲突
lock = Lock()

def get_qwq_response(messages, local_flag=False):
    """
    根据 local_flag 调用本地 Qwen/QwQ-32B 或者 DashScope API。
    """
    if local_flag:
        model_name = "Qwen/Qwen3-235B-A22B"
        model = AutoModelForCausalLM.from_pretrained(
            model_name,
            torch_dtype="auto",
            device_map="auto"
        )
        tokenizer = AutoTokenizer.from_pretrained(model_name)
        
        text = tokenizer.apply_chat_template(
            messages,
            tokenize=False,
            add_generation_prompt=True
        )
        model_inputs = tokenizer([text], return_tensors="pt").to(model.device)
        generated_ids = model.generate(**model_inputs, max_new_tokens=32768)
        # 截取新增的 tokens
        generated_ids = [
            output_ids[len(input_ids):] 
            for input_ids, output_ids in zip(model_inputs.input_ids, generated_ids)
        ]
        response = tokenizer.batch_decode(generated_ids, skip_special_tokens=True)[0]
        answer = response.lower()
        return answer
    else:
        max_retries = 10
        retry_delay = 5
        for attempt in range(max_retries):
            try:
                response = client.chat.completions.create(
                    model="qwq-32b",
                    messages=messages,
                    temperature=0,
                    stream=True, 
                    timeout=60,
                )
                break
            except:
                print(f"[Retrying] Internal server error on attempt {attempt+1}/{max_retries}. Waiting {retry_delay}s...")
                time.sleep(retry_delay)
                retry_delay *= 2
        else:
            print("[Exception] Maximum retries reached. Skipping this record.")
            return "Exception：Internal Server Error"
        
        answer_content = ""
        is_answering = False
        try:
            for chunk in response:
                delta = chunk.choices[0].delta
                # 跳过思考过程
                if hasattr(delta, 'reasoning_content') and delta.reasoning_content is not None:
                    continue
                else:
                    if delta.content and not is_answering:
                        is_answering = True
                    answer_content += delta.content
            answer = answer_content.lower()
        except:
            print("[Exception] Chunk error. Skipping this record.")
            return "Exception：Internal Server Error"

        return answer

def read_abstract_dicts(file_path: str):
    """
    使用 pandas 读取 CSV 文件后，将每一条记录转换成字典并返回 List[dict]。
    """
    df = pd.read_csv(file_path, encoding='utf-8', encoding_errors='replace')
    return df.to_dict(orient='records')

def read_method_dicts(folder_path):
    result = {}
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        
        # 确保是文件而不是子文件夹
        if os.path.isfile(file_path):
            name_without_ext = os.path.splitext(filename)[0]  # 去掉后缀
            
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
                
            result[name_without_ext] = content
    return result

def combined_abstract_method(records_abstract, records_method):
    new_list = []
    for record_dic in records_abstract:
        if str(record_dic["PMID"]) in records_method.keys():
            method = records_method[str(record_dic["PMID"])]
            record_dic["Method"] = method
            new_list.append(record_dic)
    return new_list

def judge_record(df_filtered, record):
    """
    调用 QwQ 模型判断记录是 Include 或 Exclude。
    """

    messages, screening_criteria, last_exclusion_index = return_messages(df_filtered, record)

    answer = get_qwq_response(messages, local_flag=False)

    # 解析答案
    ans_pattern = r"<answer>\s*(.*?)\s*</answer>"
    reason_pattern = r"<reasons>\s*(.*?)\s*</reasons>"
    ambiguity_pattern = r"<ambiguity>\s*(.*?)\s*</ambiguity>"

    try:
        answer_res = re.findall(ans_pattern, answer, re.DOTALL)[-1]
    except:
        # 如果无法解析 answer 或 reason
        return ("Exception", "Re Exception  \n"+ answer, "Re Exception  \n"+answer, "Re Exception  \n", screening_criteria)
    
    try:
        reason_res = re.findall(reason_pattern, answer, re.DOTALL)[-1]
    except:
        reason_res = "NA"

    try:
        ambiguity_res = re.findall(ambiguity_pattern, answer, re.DOTALL)
        num_extract = re.findall(r"ambiguity at criterion (\d+)", str(ambiguity_res), re.DOTALL)
        ambiguity_res = "Ambiguity at "+ str([int(i) for i in num_extract])
    except:
        ambiguity_res = "NA"

    # 判断包含 Include
    if "include" in answer_res.lower():
        return ("Include", reason_res, answer, ambiguity_res, screening_criteria)

    # 判断 Excluded at Criterion X
    for i in range(1, last_exclusion_index + 1):
        exclusion_str = f"excluded at criterion {i}"
        if exclusion_str in answer_res.lower():
            return (f"Excluded at Criterion {i}", reason_res, answer, ambiguity_res, screening_criteria)

    pattern = re.compile(r'(?i)(excluded at criterion)[^0-9\n]*([0-9]+)[^0-9\n]*')
    for i in range(1, last_exclusion_index + 1):
        if f"excluded at criterion {i}" in pattern.sub(r'\1 \2', answer_res.lower()):
            return (f"Excluded at Criterion {i}", reason_res, answer, ambiguity_res, screening_criteria)
        
    # 否则返回 Exception
    return ("Exception", answer, answer, ambiguity_res, screening_criteria)


def save_results_to_excel_old(results, output_path):
    """
    将本组记录结果写入 Excel（追加模式），按 Decision 分 sheet。
    加锁避免多个线程同时写。
    """
    df = pd.DataFrame(results)
    if df.empty:
        return

    with lock:
        if not os.path.exists(output_path):
            # 新建文件
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                for decision in df["Decision"].unique():
                    df_subset = df[df["Decision"] == decision]
                    df_subset.to_excel(writer, sheet_name=decision, index=False)
        else:
            # 追加模式
            with pd.ExcelWriter(output_path, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
                book = writer.book
                for decision in df["Decision"].unique():
                    df_subset = df[df["Decision"] == decision]
                    if decision not in book.sheetnames:
                        df_subset.to_excel(writer, sheet_name=decision, index=False)
                    else:
                        sheet = book[decision]
                        start_row = sheet.max_row
                        df_subset.to_excel(writer, sheet_name=decision, index=False,
                                           header=False, startrow=start_row)


def save_results_to_excel(results, output_file):
    """
    将一批 results 追加写入 Excel（按 Decision 分 sheet）。
    先检查 Exception sheet 中是否有重复 PMID，若有则删除旧记录，
    然后再将本次 results 按 Decision 写入对应 sheet 中（如果 Exception sheet 被删也会重建）。
    """
    if not results:
        return

    df = pd.DataFrame(results)
    pmid_list = set(df["PMID"].tolist())

    with lock:
        # —— 1. 准备或加载工作簿 —— #
        if not os.path.exists(output_file):
            wb = Workbook()
            wb.remove(wb.active)  # 删掉默认的空 sheet
        else:
            wb = load_workbook(output_file)
            # —— 2. 清理 Exception sheet 中重复 PMID —— #
            if "Exception" in wb.sheetnames:
                df_exc = pd.read_excel(output_file, sheet_name="Exception")
                df_exc_filtered = df_exc[~df_exc["PMID"].isin(pmid_list)]
                wb.remove(wb["Exception"])
                # 重建 Exception sheet 并写回旧数据
                ws_exc = wb.create_sheet("Exception")
                for col_idx, col_name in enumerate(df_exc_filtered.columns, start=1):
                    ws_exc.cell(row=1, column=col_idx, value=col_name)
                for row in df_exc_filtered.itertuples(index=False):
                    ws_exc.append(row)

        # —— 3. 将本次 results 按 Decision 分 sheet 写入 —— #
        for decision_val, group_df in df.groupby("Decision"):
            sheet_name = str(decision_val)

            # 如果这个 decision 是 "Exception"，或者 sheet 被删掉，都需要先确保它存在
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                # 写 header
                for col_idx, col_name in enumerate(group_df.columns, start=1):
                    ws.cell(row=1, column=col_idx, value=col_name)
                # 写数据
                for row in group_df.itertuples(index=False):
                    ws.append(row)
            else:
                # 已存在则直接 append 数据（不写 header）
                ws = wb[sheet_name]
                for row in group_df.itertuples(index=False):
                    ws.append(row)

        if "Exception" in wb.sheetnames:
            ws_exc = wb["Exception"]
            # ws_exc.max_row == 0: 完全空； ==1: 只有表头
            if ws_exc.max_row <= 1:
                wb.remove(ws_exc)

        # —— 4. 保存 —— #
        wb.save(output_file)




def process_dataframe_by_id(df, target_id):
    """
    从 df 中取出指定 ID 的行，并解析成 {Objective, Inclusion_Criteria, Exclusion_Criteria}。
    """
    df_filtered = df[df['ID'] == target_id].reset_index(drop=True)
    if df_filtered.empty:
        raise ValueError(f"No rows found for ID={target_id}.")

    if df_filtered.loc[0, 'Type'] != 'Objective':
        raise ValueError("The first row does not have Type = 'Objective'.")

    result_dict = {"Objective": df_filtered.loc[0, 'Criteria']}
    inc_criteria, inc_types = [], []
    exc_criteria, exc_types = [], []

    for _, row in df_filtered.iloc[1:].iterrows():
        if row['Abstract_Screening'] =='Yes':    
            if row['Type'] == 'Inclusion':
                inc_criteria.append(row['Criteria'])
                inc_types.append(row['Category'])
            elif row['Type'] == 'Exclusion':
                exc_criteria.append(row['Criteria'])
                exc_types.append(row['Category'])
            else:
                raise ValueError(f"Unexpected Type: {row['Type']}")

    result_dict.update({
        "Inclusion_Criteria": inc_criteria,
        "Inclusion_Type": inc_types,
        "Exclusion_Criteria": exc_criteria,
        "Exclusion_Type": exc_types,
    })
    return result_dict


def filter_criteria_by_type(data_dict, need_list):
    """
    只保留 data_dict 中类型在 need_list 里的条目。
    """
    # Inclusion
    if "Inclusion_Criteria" in data_dict and "Inclusion_Type" in data_dict:
        new_incl = []
        new_incl_t = []
        for c, t in zip(data_dict["Inclusion_Criteria"], data_dict["Inclusion_Type"]):
            if t in need_list:
                new_incl.append(c)
                new_incl_t.append(t)
        data_dict["Inclusion_Criteria"] = new_incl
        data_dict["Inclusion_Type"] = new_incl_t
    else:
        raise KeyError("Missing 'Inclusion_Criteria' or 'Inclusion_Type' in data_dict.")

    # Exclusion
    if "Exclusion_Criteria" in data_dict and "Exclusion_Type" in data_dict:
        new_exc = []
        new_exc_t = []
        for c, t in zip(data_dict["Exclusion_Criteria"], data_dict["Exclusion_Type"]):
            if t in need_list:
                new_exc.append(c)
                new_exc_t.append(t)
        data_dict["Exclusion_Criteria"] = new_exc
        data_dict["Exclusion_Type"] = new_exc_t
    else:
        # 如果没有 Exclusion，不报错
        pass

    return data_dict


def process_one_record(index, rec, filtered_dict):
    """
    处理单条记录，返回结果 dict。
    """
    decision, reason, response_text, ambiguity, prompt = judge_record(filtered_dict, rec)
    
    return {
        "Index": index,
        "PMID": rec.get("PMID", ""),
        "Title": rec.get("Title", ""),
        "Abstract": rec.get("Abstract", ""),
        "Method": rec.get("Method", ""),
        "Decision": decision,
        "Ambiguity": ambiguity,
        "Reason": reason,
        "Prompt": prompt,
        "Response": response_text,
    }


def process_one_group(records_group, filtered_dict, start_index, output_file):
    """
    处理一组记录。处理完成后一次性写入文件。
    """
    group_results = []
    for idx, rec in enumerate(records_group):
        real_idx = start_index + idx
        one_res = process_one_record(real_idx, rec, filtered_dict)
        group_results.append(one_res)

    # 当前组处理完，统一写文件
    save_results_to_excel(group_results, output_file)
    return group_results


def qwq_main(tmp_dic=None):
    num_groups = 50
    if len(sys.argv) < 3 and tmp_dic==None:
        file_ID = 490
        need_ids_dic = {}
        abstract_file_path = f"data/abstract_title_test/{file_ID}_pubmed_results.csv"
        method_file_path = f"data/OCR_extracted_test/SR_{file_ID}_test"
        print("==testing==========================================")
    elif len(sys.argv) < 3 and len(tmp_dic)>0:
        file_ID = list(tmp_dic.keys())[0]
        need_ids_dic=tmp_dic
        abstract_file_path = f"data/abstract_title/{file_ID}_pubmed_results.csv"
        method_file_path = f"data/OCR_extracted/SR_{file_ID}"
    else:
        file_ID = int(os.path.basename(sys.argv[1]).split('_')[0])  
        need_ids_dic = eval(sys.argv[2])
        abstract_file_path = f"data/abstract_title/{file_ID}_pubmed_results.csv"
        method_file_path = f"data/OCR_extracted/SR_{file_ID}"

    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    abstract_file_path = os.path.join(base_dir, abstract_file_path)
    method_file_path = os.path.join(base_dir, method_file_path)


    sr_file_path = os.path.join(base_dir, 'crawler_download', 'Data Abstraction.xlsx')
    df_criteria = pd.read_excel(sr_file_path, sheet_name="Criteria")

    res_folder_path = "step2/res/qwq_results_multi"
    res_folder_path = os.path.join(base_dir, res_folder_path)
    if not os.path.exists(res_folder_path):
        os.makedirs(res_folder_path)
    
    output_file = f"{res_folder_path}/{file_ID}_results.xlsx"
    output_file = os.path.join(base_dir, output_file)
    
    if os.path.exists(output_file) and (len(need_ids_dic)==0):
        os.remove(output_file)

    records_abstract = read_abstract_dicts(abstract_file_path)
    records_method = read_method_dicts(method_file_path)
    records_combined = combined_abstract_method(records_abstract, records_method)

    processed_data_dict = process_dataframe_by_id(df_criteria, file_ID)
    filtered_result = filter_criteria_by_type(processed_data_dict, ["Participant", "Intervention", "Control"])

    # 跳过已处理的部分
    if file_ID  in need_ids_dic.keys():
        rec_subset = []
        for item in records_combined:
            if item['PMID'] in need_ids_dic[file_ID]:
                rec_subset.append(item)
    else:
        rec_subset = records_combined

    total_len = len(rec_subset)
    if total_len == 0:
        raise Exception(f"[Info] file {file_ID} => no new records to process.")

    total_len = len(rec_subset)
    chunk_size = math.ceil(total_len / num_groups)
    groups = []
    for i in range(0, total_len, chunk_size):
        groups.append(rec_subset[i : i + chunk_size])

    print(f"[Info] file={file_ID}, total={total_len}, splitted into {len(groups)} groups")

    from concurrent.futures import ThreadPoolExecutor
    all_results = []
    finish_flag = True
    all_failed_index = []
    with ThreadPoolExecutor(max_workers=num_groups) as executor:
        futures = {}
        for g_idx, g_data in enumerate(groups):
            # 每组的起始 index
            group_start_idx =  g_idx*chunk_size
            fut = executor.submit(process_one_group, g_data, filtered_result, group_start_idx, output_file)
            futures[fut] = g_idx

        for fut in as_completed(futures):
            g_idx = futures[fut]
            try:
                partial_results = fut.result()
                all_results.extend(partial_results)
                # print(f"[Success] group-{g_idx} done, {len(partial_results)} records written.")
            except Exception as e:
                finish_flag=False
                failed_list = [i['PMID'] for i in g_data]
                all_failed_index.extend(failed_list)
    
    if finish_flag:
        print(f"[Info] file_ID={file_ID} => all {len(groups)} groups finished.")
    else:
        print(f"[Error] file_ID={file_ID} =>  {all_failed_index}  failed.")


def my_function(n):
    # 模拟要被测量的代码
    s = 0
    for i in range(n):
        s += i
    return s

if __name__ == "__main__":
    start_time = time.time()        # 记录开始时间
    qwq_main()
    end_time = time.time() 
    elapsed = end_time - start_time
    minutes = int(elapsed // 60)
    seconds = elapsed % 60
    print(f"my_function 执行时间：{minutes} min {seconds:.2f} second \n\n")