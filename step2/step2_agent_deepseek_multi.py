import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import re
import time
import math
import json
import openpyxl
import pandas as pd
from openpyxl import Workbook
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
from openai import OpenAI
from prompts.step2_common_0 import return_messages
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()  
api_key = os.getenv("ALI_API_KEY")
client = OpenAI(
    api_key=api_key,  
    base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",  # DashScope
)
model_type="deepseek-r1"

# 线程锁，用于并发写 Excel 文件
lock = Lock()

def read_abstract_dicts(file_path: str):
    """
    使用 pandas 读取 CSV 文件后，将每一条记录转换成字典并返回 List[dict]。
    """
    df = pd.read_csv(file_path, encoding='utf-8', encoding_errors='replace')
    return df.to_dict(orient='records')

def read_method_dicts(folder_path):
    result = {}
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        
        # 确保是文件而不是子文件夹
        if os.path.isfile(file_path):
            name_without_ext = os.path.splitext(filename)[0]  # 去掉后缀
            
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
                
            result[name_without_ext] = content
    return result

def combined_abstract_method(records_abstract, records_method):
    new_list = []
    for record_dic in records_abstract:
        if str(record_dic["PMID"]) in records_method.keys():
            method = records_method[str(record_dic["PMID"])]
            record_dic["Method"] = method
            new_list.append(record_dic)
    return new_list

def judge_record(g_idx, df_filtered, record, output_file):
    """
    调用 deepseek-r1 模型判断该记录的纳排情况。
    """


    messages, screening_criteria, last_exclusion_index = return_messages(df_filtered, record)
    output_file = output_file.split("/")[-1]

    # 重试机制
    max_retries = 10
    retry_delay = 5
    response = None
    for attempt in range(max_retries):
        try:
            response = client.chat.completions.create(
                model=model_type,
                messages=messages,
                temperature=0,
                stream=True, 
                timeout=60
            )
            break
        except Exception as e:
            print(f"[Retrying] {output_file}:{g_idx} Internal server error on attempt {attempt+1}/{max_retries}. error={e}. Waiting {retry_delay}s...")
            time.sleep(retry_delay)
            retry_delay *= 2
    
    if not response:
        print("[Exception] {output_file}:{g_idx} Maximum retries reached. Skipping this record.")
        return ("Exception", "Internal Server Error", "NA", "NA", screening_criteria)

    # 解析
    answer_content = ""
    is_answering = False
    reasoning_content = ""
    try:
        for chunk in response:
            delta = chunk.choices[0].delta
            if hasattr(delta, 'reasoning_content') and delta.reasoning_content is not None:
                reasoning_content += delta.reasoning_content
            else:
                if delta.content and not is_answering:
                    is_answering = True
                answer_content += delta.content
    except:
        print("[Exception] Chunk error. Skipping this record.")
        return ("Exception", "Chunk error", "NA", "NA", screening_criteria)
    
    answer_text = answer_content.strip().lower()
    ans_pattern = r"<answer>\s*(.*?)\s*</answer>"
    reason_pattern = r"<reasons>\s*(.*?)\s*</reasons>"
    ambiguity_pattern = r"<ambiguity>\s*(.*?)\s*</ambiguity>"

    try:
        answer_res = re.findall(ans_pattern, answer_text, re.DOTALL)[-1]
    except:
        return ("Exception", "!!! Re Exception !!!  \n"+reasoning_content, "!!! Re Exception !!!  \n"+answer_text, "!!! Re Exception !!!  \n", screening_criteria)

    reason_res = reasoning_content
  
    try:
        ambiguity_res = re.findall(ambiguity_pattern, answer_text, re.DOTALL)
        num_extract = re.findall(r"ambiguity at criterion (\d+)", str(ambiguity_res), re.DOTALL)
        ambiguity_res = "Ambiguity at "+ str([int(i) for i in num_extract])
    except:
        ambiguity_res = "NA"

    if "included" in answer_res.lower():
        return ("Include", reason_res, answer_text, ambiguity_res, screening_criteria)

    pattern = re.compile(r'(?i)(excluded at criterion)[^0-9\n]*([0-9]+)[^0-9\n]*')
    for i in range(1, last_exclusion_index + 1):
        if f"excluded at criterion {i}" in pattern.sub(r'\1 \2', answer_res.lower()):
            return (f"Excluded at Criterion {i}", reason_res, answer_text, ambiguity_res, screening_criteria)

    return ("Exception", "!!! Final Exception !!!  \n"+ answer_text, "!!! Final Exception !!!  \n"+ answer_text, "!!! Final Exception !!!  \n"+ ambiguity_res, screening_criteria)
              

def append_results_to_excel_old(results, output_path):
    """
    将一批记录的结果追加到 Excel（按 Decision 分 sheet）。
    用 lock 保证并发写安全。
    """
    if not results:
        return
    
    df_all = pd.DataFrame(results)
    with lock:
        if not os.path.exists(output_path):
            # 新建文件
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                for decision_val, group_df in df_all.groupby("Decision"):
                    group_df.to_excel(writer, sheet_name=str(decision_val), index=False)
        else:
            # 追加模式
            with pd.ExcelWriter(output_path, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
                book = writer.book
                for decision_val, group_df in df_all.groupby("Decision"):
                    sheet_name = str(decision_val)
                    if sheet_name not in book.sheetnames:
                        group_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    else:
                        sheet = book[sheet_name]
                        start_row = sheet.max_row
                        group_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=start_row)


def append_results_to_excel(results, output_file):
    """
    将一批 results 追加写入 Excel（按 Decision 分 sheet）。
    先检查 Exception sheet 中是否有重复 PMID，若有则删除旧记录，
    然后再将本次 results 按 Decision 写入对应 sheet 中（如果 Exception sheet 被删也会重建）。
    """
    if not results:
        return

    df = pd.DataFrame(results)
    pmid_list = set(df["PMID"].tolist())

    with lock:
        # —— 1. 准备或加载工作簿 —— #
        if not os.path.exists(output_file):
            wb = Workbook()
            wb.remove(wb.active)  # 删掉默认的空 sheet
        else:
            wb = load_workbook(output_file)
            # —— 2. 清理 Exception sheet 中重复 PMID —— #
            if "Exception" in wb.sheetnames:
                df_exc = pd.read_excel(output_file, sheet_name="Exception")
                df_exc_filtered = df_exc[~df_exc["PMID"].isin(pmid_list)]
                wb.remove(wb["Exception"])
                # 重建 Exception sheet 并写回旧数据
                ws_exc = wb.create_sheet("Exception")
                for col_idx, col_name in enumerate(df_exc_filtered.columns, start=1):
                    ws_exc.cell(row=1, column=col_idx, value=col_name)
                for row in df_exc_filtered.itertuples(index=False):
                    ws_exc.append(row)

        # —— 3. 将本次 results 按 Decision 分 sheet 写入 —— #
        for decision_val, group_df in df.groupby("Decision"):
            sheet_name = str(decision_val)

            # 如果这个 decision 是 "Exception"，或者 sheet 被删掉，都需要先确保它存在
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                # 写 header
                for col_idx, col_name in enumerate(group_df.columns, start=1):
                    ws.cell(row=1, column=col_idx, value=col_name)
                # 写数据
                for row in group_df.itertuples(index=False):
                    ws.append(row)
            else:
                # 已存在则直接 append 数据（不写 header）
                ws = wb[sheet_name]
                for row in group_df.itertuples(index=False):
                    ws.append(row)

        if "Exception" in wb.sheetnames:
            ws_exc = wb["Exception"]
            # ws_exc.max_row == 0: 完全空； ==1: 只有表头
            if ws_exc.max_row <= 1:
                wb.remove(ws_exc)

        # —— 4. 保存 —— #
        wb.save(output_file)




def process_dataframe_by_id(df, target_id):
    df_filtered = df[df['ID'] == target_id].reset_index(drop=True)
    if df_filtered.empty:
        raise ValueError(f"No rows found for ID={target_id}.")

    if df_filtered.loc[0, 'Type'] != 'Objective':
        raise ValueError("The first row does not have Type = 'Objective'.")

    result_dict = {"Objective": df_filtered.loc[0, 'Criteria']}
    inclusion_criteria, inclusion_types = [], []
    exclusion_criteria, exclusion_types = [], []

    for _, row in df_filtered.iloc[1:].iterrows():
        if row['Abstract_Screening'] =='Yes':
            if row['Type'] == 'Inclusion':
                inclusion_criteria.append(row['Criteria'])
                inclusion_types.append(row['Category'])
            elif row['Type'] == 'Exclusion':
                exclusion_criteria.append(row['Criteria'])
                exclusion_types.append(row['Category'])
            else:
                raise ValueError(f"Unexpected Type={row['Type']}")

    result_dict.update({
        "Inclusion_Criteria": inclusion_criteria,
        "Inclusion_Type": inclusion_types,
        "Exclusion_Criteria": exclusion_criteria,
        "Exclusion_Type": exclusion_types
    })
    return result_dict


def filter_criteria_by_type(data_dict, need_list):
    """
    只保留 'Inclusion_Criteria' / 'Exclusion_Criteria' 中类型属于 need_list 的部分
    """
    # Inclusion
    if "Inclusion_Criteria" in data_dict and "Inclusion_Type" in data_dict:
        new_incl_criteria = []
        new_incl_type = []
        for c, t in zip(data_dict["Inclusion_Criteria"], data_dict["Inclusion_Type"]):
            if t in need_list:
                new_incl_criteria.append(c)
                new_incl_type.append(t)
        data_dict["Inclusion_Criteria"] = new_incl_criteria
        data_dict["Inclusion_Type"] = new_incl_type
    else:
        raise KeyError("Missing keys: 'Inclusion_Criteria' and/or 'Inclusion_Type'.")

    # Exclusion
    if "Exclusion_Criteria" in data_dict and "Exclusion_Type" in data_dict:
        new_excl_criteria = []
        new_excl_type = []
        for c, t in zip(data_dict["Exclusion_Criteria"], data_dict["Exclusion_Type"]):
            if t in need_list:
                new_excl_criteria.append(c)
                new_excl_type.append(t)
        data_dict["Exclusion_Criteria"] = new_excl_criteria
        data_dict["Exclusion_Type"] = new_excl_type

    return data_dict


def process_one_record(g_idx, index, rec, filtered_result, output_file):
    """
    单条记录处理
    """
    decision, reason, response, ambiguity, screening_text = judge_record(g_idx, filtered_result, rec, output_file)
    
    return {
        "Index": index,
        "PMID": rec.get("PMID", ""),
        "Title": rec.get("Title", ""),
        "Abstract": rec.get("Abstract", ""),
        "Method": rec.get("Method", ""),
        "Decision": decision,
        "Ambiguity": ambiguity,
        "Reason": reason,
        "Prompt": screening_text,
        "Response": response,
    }

    

def process_one_group(g_idx, records_in_group, filtered_result, start_index, output_file):
    """
    并发时：处理一组记录，然后把结果一次性写到 Excel。
    """
    group_results = []
    for i, rec in enumerate(records_in_group):
        real_idx = start_index + i
        res = process_one_record(g_idx, real_idx, rec, filtered_result, output_file)
        group_results.append(res)

    # 处理完该组后，一次性追加写文件
    append_results_to_excel(group_results, output_file)
    return group_results


def deepseek_main(tmp_dic=None):
    num_groups = 30
    if len(sys.argv) < 3 and tmp_dic==None:
        file_ID = 490
        need_ids_dic = {}
        abstract_file_path = f"data/abstract_title_test/{file_ID}_pubmed_results.csv"
        method_file_path = f"data/OCR_extracted_test/SR_{file_ID}_test"
        print("==testing==========================================")
    elif len(sys.argv) < 3 and len(tmp_dic)>0:
        file_ID = list(tmp_dic.keys())[0]
        need_ids_dic=tmp_dic
        abstract_file_path = f"data/abstract_title/{file_ID}_pubmed_results.csv"
        method_file_path = f"data/OCR_extracted/SR_{file_ID}"
    else:
        file_ID = int(os.path.basename(sys.argv[1]).split('_')[0])  
        need_ids_dic = eval(sys.argv[2])
        abstract_file_path = f"data/abstract_title/{file_ID}_pubmed_results.csv"
        method_file_path = f"data/OCR_extracted/SR_{file_ID}"


    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    abstract_file_path = os.path.join(base_dir, abstract_file_path)
    method_file_path = os.path.join(base_dir, method_file_path)

    # 读取筛选标准 Excel
    sr_file_path = os.path.join(base_dir, 'crawler_download', 'Data Abstraction.xlsx')
    df_criteria = pd.read_excel(sr_file_path, sheet_name="Criteria")

    res_folder_path = "step2/res/deepseek_results_multi"
    res_folder_path = os.path.join(base_dir, res_folder_path)
    if not os.path.exists(res_folder_path):
        os.makedirs(res_folder_path)

    output_file = f"{res_folder_path}/{file_ID}_results.xlsx"
    output_file = os.path.join(base_dir, output_file)

    if os.path.exists(output_file) and (len(need_ids_dic)==0):
        os.remove(output_file)

    records_abstract = read_abstract_dicts(abstract_file_path)
    records_method = read_method_dicts(method_file_path)
    records_combined = combined_abstract_method(records_abstract, records_method)

    processed_data_dict = process_dataframe_by_id(df_criteria, file_ID)
    filtered_result = filter_criteria_by_type(processed_data_dict, ["Participant", "Intervention", "Control"])

    # 跳过已处理的部分
    if file_ID  in need_ids_dic.keys():
        rec_subset = []
        for item in records_combined:
            if item['PMID'] in need_ids_dic[file_ID]:
                rec_subset.append(item)
    else:
        rec_subset = records_combined

    total_len = len(rec_subset)
    if total_len == 0:
        raise Exception(f"[Info] file {file_ID} => no new records to process.")

    total_len = len(rec_subset)
    chunk_size = math.ceil(total_len / num_groups)
    groups = []
    for i in range(0, total_len, chunk_size):
        groups.append(rec_subset[i : i + chunk_size])

    print(f"[Info] file={file_ID}, total={total_len}, splitted into {len(groups)} groups")

    from concurrent.futures import ThreadPoolExecutor
    all_results = []
    finish_flag = True
    all_failed_index = []
    with ThreadPoolExecutor(max_workers=num_groups) as executor:
        futures = {}
        for g_idx, g_data in enumerate(groups):
            # 每组的起始 index
            group_start_idx =  g_idx*chunk_size
            fut = executor.submit(process_one_group, g_idx, g_data, filtered_result, group_start_idx, output_file)
            futures[fut] = g_idx

        for fut in as_completed(futures):
            g_idx = futures[fut]
            try:
                partial_results = fut.result()
                all_results.extend(partial_results)
                # print(f"[Success] group-{g_idx} done, {len(partial_results)} records written.")
            except Exception as e:
                finish_flag=False
                failed_list = [i['PMID'] for i in g_data]
                all_failed_index.extend(failed_list)

    if finish_flag:
        print(f"[Info] file_ID={file_ID} => all {len(groups)} groups finished.")
    else:
        print(f"[Error] file_ID={file_ID} =>  {all_failed_index}  failed.")


if __name__ == "__main__":
    start_time = time.time()        # 记录开始时间
    deepseek_main()
    end_time = time.time() 
    elapsed = end_time - start_time
    minutes = int(elapsed // 60)
    seconds = elapsed % 60
    print(f"my_function 执行时间：{minutes} min {seconds:.2f} second \n\n")