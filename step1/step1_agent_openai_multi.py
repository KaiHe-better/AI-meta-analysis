import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import math
import time
import re
import json
import openai
import pandas as pd
import openpyxl
from openpyxl import Workbook
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock  # 用于并发写文件加锁
from prompts.step1_common_42 import return_messages
from openpyxl import load_workbook
from dotenv import load_dotenv


load_dotenv()  # 加载 .env 文件
api_key = os.getenv("OPENAI_API_KEY")
client = openai.OpenAI(api_key=api_key)

# 全局锁，用于串行化写文件操作
lock = Lock()

def read_csv_return_dicts(file_path: str):
    """
    使用 pandas 读取 CSV 文件后，将每一条记录转换成字典并返回 List[dict]。
    """
    df = pd.read_csv(file_path, encoding='utf-8', encoding_errors='replace')
    return df.to_dict(orient='records')

# o3  o4-mini
def judge_record(df_filtered, record: dict, model: str = "o3"):
    """
    调用 OpenAI ChatCompletion API，判断该记录是 "Include" 还是 "Exclude"。
    返回 (decision, reason, ambiguity, screening_criteria) 四元组。
    """

    messages, screening_criteria, last_exclusion_index = return_messages(df_filtered, record)

    max_retries = 5
    retry_delay = 5
    response_text = None
    for attempt in range(max_retries):
        try:
            response = client.chat.completions.create(
                model=model,
                messages=messages,
                timeout=60,
            )
            response_text = response.choices[0].message.content.strip().lower()
            break
        except Exception as e:
            print(f"[Retry] attempt {attempt+1}/{max_retries}, error: {e}")
            time.sleep(retry_delay)
            retry_delay *= 2
 
    if not response_text:
        # 如果最终都没拿到结果
        return ("Exception", "Network Exception", "Network Exception", screening_criteria)

    # 解析返回文本
    anwser_pattern = r"<answer>\s*(.*?)\s*</answer>"
    reason_pattern = r"<reasons>\s*(.*?)\s*</reasons>"
    ambiguity_pattern = r"<ambiguity>\s*(.*?)\s*</ambiguity>"

    try:
        answer_res = re.findall(anwser_pattern, response_text, re.DOTALL)[-1]
    except:
        return ("Exception", "Re Exception \n"+ response_text, "Re Exception  \n"+response_text, "Re Exception  \n", screening_criteria)

    try:
        reason_res = re.findall(reason_pattern, response_text, re.DOTALL)[-1]
    except:
        reason_res = "NA"

    try:
        ambiguity_res = re.findall(ambiguity_pattern, response_text, re.DOTALL)
        num_extract = re.findall(r"ambiguity at criterion (\d+)", str(ambiguity_res), re.DOTALL)
        ambiguity_res = "Ambiguity at "+ str([int(i) for i in num_extract])
    except:
        ambiguity_res = "NA"

    # 判断结果
    if "included" in answer_res.lower():
        return ("Include", reason_res, response_text, ambiguity_res, screening_criteria)
    
    pattern = re.compile(r'(?i)(excluded at criterion)[^0-9\n]*([0-9]+)[^0-9\n]*')
    for i in range(1, last_exclusion_index + 1):
        if f"excluded at criterion {i}" in pattern.sub(r'\1 \2', answer_res.lower()):
            return (f"Excluded at Criterion {i}", reason_res, response_text, ambiguity_res, screening_criteria)

    return ("Exception", "Final Exception  \n"+ response_text, "Final Exception  \n"+response_text, "Final Exception  \n"+ambiguity_res, screening_criteria)


def process_dataframe_by_id(df, target_id):
    """
    从excel中取出某个ID的行，并解析得到Objective、Inclusion/Exclusion。
    """
    df_filtered = df[df['ID'] == target_id].reset_index(drop=True)
    if df_filtered.empty:
        raise ValueError(f"No records found for ID={target_id}.")

    if df_filtered.loc[0, 'Type'] != 'Objective':
        raise ValueError("The first row is not Type='Objective'!")

    result_dict = {"Objective": df_filtered.loc[0, 'Criteria']}
    inclusion_criteria, inclusion_types = [], []
    exclusion_criteria, exclusion_types = [], []

    for _, row in df_filtered.iloc[1:].iterrows():
        if row['Abstract_Screening'] =='Yes':
            if row['Type'] == 'Inclusion':
                inclusion_criteria.append(row['Criteria'])
                inclusion_types.append(row['Category'])
            elif row['Type'] == 'Exclusion':
                exclusion_criteria.append(row['Criteria'])
                exclusion_types.append(row['Category'])
            else:
                raise ValueError(f"Unexpected Type found: {row['Type']}")

    result_dict.update({
        "Inclusion_Criteria": inclusion_criteria,
        "Inclusion_Type": inclusion_types,
        "Exclusion_Criteria": exclusion_criteria,
        "Exclusion_Type": exclusion_types
    })
    return result_dict


def filter_criteria_by_type(data_dict, need_list):
    """
    只保留 need_list 类型的 Inclusion/Exclusion
    """
    # Inclusion
    if "Inclusion_Criteria" in data_dict and "Inclusion_Type" in data_dict:
        new_incl_criteria = []
        new_incl_types = []
        for c, t in zip(data_dict["Inclusion_Criteria"], data_dict["Inclusion_Type"]):
            if t in need_list:
                new_incl_criteria.append(c)
                new_incl_types.append(t)
        data_dict["Inclusion_Criteria"] = new_incl_criteria
        data_dict["Inclusion_Type"] = new_incl_types
    else:
        raise KeyError("Missing keys: 'Inclusion_Criteria' or 'Inclusion_Type'.")

    # Exclusion
    if "Exclusion_Criteria" in data_dict and "Exclusion_Type" in data_dict:
        new_excl_criteria = []
        new_excl_types = []
        for c, t in zip(data_dict["Exclusion_Criteria"], data_dict["Exclusion_Type"]):
            if t in need_list:
                new_excl_criteria.append(c)
                new_excl_types.append(t)
        data_dict["Exclusion_Criteria"] = new_excl_criteria
        data_dict["Exclusion_Type"] = new_excl_types

    return data_dict


def process_one_record(index, rec, filtered_result):
    """
    单条记录处理逻辑
    """
    decision, reason, response_text, ambiguity_res, prompt = judge_record(filtered_result, rec)
    return {
        "Index": index,
        "PMID": rec.get("PMID", ""),
        "Title": rec.get("Title", ""),
        "Abstract": rec.get("Abstract", ""),
        "Decision": decision,
        "Ambiguity": ambiguity_res,
        "Reason": reason,
        "Prompt": prompt,
        "Response": response_text,
    }


def append_results_to_excel_old(results, output_file):
    """
    将一批 results 追加写入 Excel（按 Decision 分 sheet）。
    使用线程锁避免多个线程同时写。
    """
    if not results:
        return

    df = pd.DataFrame(results)
    # 加锁写文件
    with lock:
        if not os.path.exists(output_file):
            # 新文件，直接按照决策分sheet写
            with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                for decision_val, group_df in df.groupby("Decision"):
                    group_df.to_excel(writer, sheet_name=str(decision_val), index=False)
        else:
            # 追加模式
            # 先打开原文件，再针对每个 decision 的sheet做 append
            with pd.ExcelWriter(output_file, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
                book = writer.book
                for decision_val, group_df in df.groupby("Decision"):
                    sheet_name = str(decision_val)
                    if sheet_name not in book.sheetnames:
                        # 如果这个sheet还不存在，创建
                        group_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    else:
                        # 如果已经存在，直接在末尾追加
                        sheet = book[sheet_name]
                        start_row = sheet.max_row
                        group_df.to_excel(writer, sheet_name=sheet_name, index=False,
                                          header=False, startrow=start_row)

def append_results_to_excel(results, output_file):
    """
    将一批 results 追加写入 Excel（按 Decision 分 sheet）。
    先检查 Exception sheet 中是否有重复 PMID，若有则删除旧记录，
    然后再将本次 results 按 Decision 写入对应 sheet 中（如果 Exception sheet 被删也会重建）。
    """
    if not results:
        return

    df = pd.DataFrame(results)
    pmid_list = set(df["PMID"].tolist())

    with lock:
        # —— 1. 准备或加载工作簿 —— #
        if not os.path.exists(output_file):
            wb = Workbook()
            wb.remove(wb.active)  # 删掉默认的空 sheet
        else:
            wb = load_workbook(output_file)
            # —— 2. 清理 Exception sheet 中重复 PMID —— #
            if "Exception" in wb.sheetnames:
                df_exc = pd.read_excel(output_file, sheet_name="Exception")
                df_exc_filtered = df_exc[~df_exc["PMID"].isin(pmid_list)]
                wb.remove(wb["Exception"])
                # 重建 Exception sheet 并写回旧数据
                ws_exc = wb.create_sheet("Exception")
                for col_idx, col_name in enumerate(df_exc_filtered.columns, start=1):
                    ws_exc.cell(row=1, column=col_idx, value=col_name)
                for row in df_exc_filtered.itertuples(index=False):
                    ws_exc.append(row)

        # —— 3. 将本次 results 按 Decision 分 sheet 写入 —— #
        for decision_val, group_df in df.groupby("Decision"):
            sheet_name = str(decision_val)

            # 如果这个 decision 是 "Exception"，或者 sheet 被删掉，都需要先确保它存在
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                # 写 header
                for col_idx, col_name in enumerate(group_df.columns, start=1):
                    ws.cell(row=1, column=col_idx, value=col_name)
                # 写数据
                for row in group_df.itertuples(index=False):
                    ws.append(row)
            else:
                # 已存在则直接 append 数据（不写 header）
                ws = wb[sheet_name]
                for row in group_df.itertuples(index=False):
                    ws.append(row)

        if "Exception" in wb.sheetnames:
            ws_exc = wb["Exception"]
            # ws_exc.max_row == 0: 完全空； ==1: 只有表头
            if ws_exc.max_row <= 1:
                wb.remove(ws_exc)

        # —— 4. 保存 —— #
        wb.save(output_file)


def process_one_group(records_in_group, filtered_result, start_index, output_file):
    """
    并发时：处理一组记录。处理结束后立即写入 Excel（减少写文件次数）。
    """
    results_group = []
    for idx, rec in enumerate(records_in_group):
        real_idx = start_index + idx
        one_res = process_one_record(real_idx, rec, filtered_result)
        results_group.append(one_res)


    append_results_to_excel(results_group, output_file)
    return results_group 


def openai_main(tmp_dic=None):
    num_groups = 50
    if len(sys.argv) < 3 and tmp_dic==None:
        file_ID = 344
        need_ids_dic = {}
        file_path = f"data/abstract_title_test/{file_ID}_pubmed_results.csv"
        print("==testing==========================================")
    elif len(sys.argv) < 3 and len(tmp_dic)>0:
        file_ID = list(tmp_dic.keys())[0]
        need_ids_dic=tmp_dic
        file_path = f"data/abstract_title/{file_ID}_pubmed_results.csv"
    else:
        file_ID = int(os.path.basename(sys.argv[1]).split('_')[0])  
        need_ids_dic = eval(sys.argv[2])
        file_path = f"data/abstract_title/{file_ID}_pubmed_results.csv"

    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    file_path = os.path.join(base_dir, file_path)
    sr_file_path = os.path.join(base_dir, 'crawler_download', 'Data Abstraction.xlsx')
    df_criteria = pd.read_excel(sr_file_path, sheet_name="Criteria")

    res_folder_path = "./res/openai_results_multi"
    if not os.path.exists(res_folder_path):
        os.makedirs(res_folder_path)
    output_file = f"./{res_folder_path}/{file_ID}_results.xlsx"
    

    if os.path.exists(output_file) and (len(need_ids_dic)==0):
        os.remove(output_file)


    records = read_csv_return_dicts(file_path)
    processed_data_dict = process_dataframe_by_id(df_criteria, file_ID)
    filtered_result = filter_criteria_by_type(processed_data_dict, ["Participant", "Intervention", "Control"])

 
    if file_ID  in need_ids_dic.keys():
        rec_subset = []
        for item in records:
            if item['PMID'] in need_ids_dic[file_ID]:
                rec_subset.append(item)
    else:
        rec_subset = records
        

    total_len = len(rec_subset)
    if total_len == 0:
        raise Exception(f"[Info] file {file_ID} => no new records to process.")

    total_len = len(rec_subset)
    chunk_size = math.ceil(total_len / num_groups)
    groups = []
    for i in range(0, total_len, chunk_size):
        groups.append(rec_subset[i : i + chunk_size])

    print(f"[Info] file={file_ID}, total={total_len}, splitted into {len(groups)} groups")

    from concurrent.futures import ThreadPoolExecutor
    all_results = []
    finish_flag = True
    all_failed_index = []
    with ThreadPoolExecutor(max_workers=num_groups) as executor:
        futures = {}
        for g_idx, g_data in enumerate(groups):
            # 每组的起始 index
            group_start_idx =  g_idx*chunk_size
            fut = executor.submit(process_one_group, g_data, filtered_result, group_start_idx, output_file)
            futures[fut] = g_idx

        for fut in as_completed(futures):
            g_idx = futures[fut]
            partial_results = fut.result()
            try:
                partial_results = fut.result()
                all_results.extend(partial_results)
                # print(f"[Success] group-{g_idx} done, {len(partial_results)} records written.")
            except Exception as e:
                finish_flag=False
                failed_list = [i['PMID'] for i in g_data]
                all_failed_index.extend(failed_list)

    if finish_flag:
        print(f"[Info] file_ID={file_ID} => all {len(groups)} groups finished.")
    else:
        print(f"[Error] file_ID={file_ID} =>  {all_failed_index}  failed.")



if __name__ == "__main__":
    start_time = time.time()        # 记录开始时间
    openai_main()
    end_time = time.time() 
    elapsed = end_time - start_time
    minutes = int(elapsed // 60)
    seconds = elapsed % 60
    print(f"my_function 执行时间：{minutes} min {seconds:.2f} second \n\n")